#!/usr/bin/env python3
"""
Extracció de dades analítiques per municipi i rang de dates.
Genera dues columnes: Ràtio Ajuntament (municipi concret) i Ràtio General (tots els municipis).

Executa:  python sacar_datos.py
"""

import sys
import calendar
from datetime import date, timedelta

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from api_client import AdtendeClient


# ─── Utilitats de data ────────────────────────────────────────────────────────

def parse_date(s: str) -> date:
    """Accepta DD/MM/YYYY o DD/MM/YY i retorna un objecte date."""
    parts = s.strip().replace("-", "/").split("/")
    if len(parts) != 3:
        raise ValueError
    d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
    if y < 100:
        y += 2000
    return date(y, m, d)


def months_in_range(d_from: date, d_to: date):
    """Llista de (any, mes) que cobreix el rang, inclusiu en els dos extrems."""
    result = []
    y, m = d_from.year, d_from.month
    while (y, m) <= (d_to.year, d_to.month):
        result.append((y, m))
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return result


def month_window(year: int, month: int):
    """(inici_str, fi_exclusiva_str) en format YYYY-MM-DD per a un mes sencer."""
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return str(start), str(end)


# ─── Lògica de descàrrega i filtre ────────────────────────────────────────────

def download_and_filter(client: AdtendeClient, d_from: date, d_to: date,
                        municipio: str = None) -> pd.DataFrame:
    """
    Descarrega mes a mes (filtre td_managed) i aplica el filtre doble.
    Si municipio=None, descarrega TOTS els municipis (Ràtio General).
    """
    s_from = str(d_from)
    s_to   = str(d_to + timedelta(days=1))

    mesos = months_in_range(d_from, d_to)
    dfs = []

    label = municipio if municipio else "GENERAL"

    for y, m in mesos:
        m_from, m_to = month_window(y, m)
        filters = [{"type": "date", "variable": "td_managed",
                    "values": {"gte": m_from, "lt": m_to}}]
        try:
            df_m = client.query("tickets_enriquits", filters=filters)
            if municipio:
                sub = df_m[
                    (df_m["des_client"]  == municipio) &
                    (df_m["des_project"] == "OAC 360º")
                ].copy()
            else:
                sub = df_m[df_m["des_project"] == "OAC 360º"].copy()
            dfs.append(sub)
            print(f"    [{label}] {y}-{m:02d}  →  {len(sub):>6} tickets")
        except Exception as e:
            print(f"    [{label}] {y}-{m:02d}  →  ERROR: {e}")

    if not dfs:
        return pd.DataFrame()

    full = pd.concat(dfs, ignore_index=True)

    df = full[
        (full["td_managed"] >= s_from) & (full["td_managed"] < s_to) &
        (full["td_created"] >= s_from) & (full["td_created"] < s_to)
    ].copy()

    return df


# ─── Càlcul dels indicadors ───────────────────────────────────────────────────

def calcular_indicadors(df: pd.DataFrame, es_general: bool = False) -> list[dict]:
    """
    Calcula els indicadors. Si es_general=True, el % Població atesa
    es calcula com la mitjana dels % per municipi (cada municipi té la seva població).
    """
    total = len(df)
    if total == 0:
        return []

    # Satisfacció — forcem numèric
    for col in ["val_rating", "val_pregunta1", "val_pregunta2", "val_pregunta3", "val_encuestable"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    def _safe_mean(series):
        m = series.mean()
        return f"{m:.2f}" if pd.notna(m) else "N/D"

    # Tots els encuestables
    enc_base = df[df["val_encuestable"] == 1]
    n_enc    = len(enc_base)

    # Subpreguntes individuals (val_encuestable==1 + col no nul·la)
    def _mean_sub(col):
        sub = enc_base[enc_base[col].notna()] if col in enc_base.columns else pd.DataFrame()
        return _safe_mean(sub[col]) if len(sub) > 0 else "N/D"
    p1 = _mean_sub("val_pregunta1")
    p2 = _mean_sub("val_pregunta2")
    p3 = _mean_sub("val_pregunta3")

    # Grau de satisfacció global: usar val_media_enc (pre-calculat a la BD, = (p1+p2+p3)/3)
    if "val_media_enc" in enc_base.columns:
        enc_base_m = enc_base.copy()
        enc_base_m["val_media_enc"] = pd.to_numeric(enc_base_m["val_media_enc"], errors="coerce")
        enc_with_enc = enc_base_m[enc_base_m["val_media_enc"].notna()]
        sat = f"{enc_with_enc['val_media_enc'].mean():.2f}" if len(enc_with_enc) > 0 else "N/D"
    else:
        # Fallback: mean(p1, p2, p3) si val_media_enc no disponible
        preg_vals = []
        for pq in ["val_pregunta1", "val_pregunta2", "val_pregunta3"]:
            if pq in enc_base.columns:
                m = enc_base[pq].mean()
                if pd.notna(m):
                    preg_vals.append(m)
        sat = f"{sum(preg_vals) / len(preg_vals):.2f}" if preg_vals else "N/D"

    # Temps mig (val_time_spent és en MINUTS) — forcem numèric
    df["val_time_spent"] = pd.to_numeric(df["val_time_spent"], errors="coerce")
    avg_t = df["val_time_spent"].mean()
    avg_m = int(avg_t) if pd.notna(avg_t) else 0
    avg_s = int((avg_t - avg_m) * 60) if pd.notna(avg_t) else 0
    temps = f"{avg_m}m {avg_s:02d}s"

    # % Població atesa
    if es_general:
        # Mitjana dels % per municipi (cada un té la seva població de referència)
        pcts = []
        for muni, grp in df.groupby("des_client"):
            pop_vals = grp["val_population"].dropna()
            if len(pop_vals) > 0:
                pop = float(pop_vals.iloc[0])
                if pop > 0:
                    pcts.append(len(grp) / pop * 100)
        pct_pop = f"{sum(pcts)/len(pcts):.2f}%" if pcts else "N/D"
        pop_str = f"{df['des_client'].nunique()} municipis"
    else:
        pop_vals = df["val_population"].dropna()
        pop      = float(pop_vals.iloc[0]) if len(pop_vals) > 0 else None
        pct_pop  = f"{total / pop * 100:.2f}%" if pop else "N/D"
        pop_str  = f"{int(pop):,}".replace(",", ".") if pop else "N/D"

    # TRUCA'M
    trucam = int(df["des_entry_channel"].isin(["TRUCA'M LOCUCIO", "TRUCA'M WEB"]).sum())

    # Franja tarda
    tarda = int((df["flg_morning_schedule"] == "15-24h").sum())

    # Cita prèvia
    cita = int((df["des_category_0"] == "Cita Prèvia").sum())

    # Catàleg de tràmits
    cat_tramits = int(df["des_category_1"].isin(["Catàleg de tràmits", "eTramitador"]).sum())

    # Campanyes puntuals
    campanyes_cats = [
        "Estat dels meus expedients", "Inscripcions", "Processos selectius",
        "eNotificacions", "Signatura electrònica", "Carpeta del Ciutadà",
    ]
    campanyes = int(df["des_category_1"].isin(campanyes_cats).sum())

    # Centraleta
    centraleta = int((df["des_category_1"] == "Trucada per centraleta").sum())

    def pct(n):
        return f"{n / total * 100:.1f}%" if total else "0%"

    return [
        {"#": "—",  "Indicador": "Total assistències",          "Valor": total,     "Pct": ""},
        {"#": "",   "Indicador": "Població (ref.)",              "Valor": pop_str,   "Pct": ""},
        {"#": "1",  "Indicador": "% Població atesa",             "Valor": pct_pop,   "Pct": ""},
        {"#": "2",  "Indicador": "Grau de satisfacció /5",       "Valor": sat,       "Pct": ""},
        {"#": "2.1","Indicador": "  Com valora la qualitat?",    "Valor": p1,        "Pct": ""},
        {"#": "2.1","Indicador": "  Li han resolt la consulta?", "Valor": p2,        "Pct": ""},
        {"#": "2.1","Indicador": "  Com valora el tracte?",      "Valor": p3,        "Pct": ""},
        {"#": "3",  "Indicador": "Temps mig consulta",           "Valor": temps,     "Pct": ""},
        {"#": "4",  "Indicador": "Ús del TRUCA'M",               "Valor": trucam,    "Pct": pct(trucam)},
        {"#": "5",  "Indicador": "Assistències franja tarda",    "Valor": tarda,     "Pct": pct(tarda)},
        {"#": "6",  "Indicador": "Assistències Cita prèvia",     "Valor": cita,      "Pct": pct(cita)},
        {"#": "7",  "Indicador": "Assistències Catàleg tràmits", "Valor": cat_tramits,"Pct": pct(cat_tramits)},
        {"#": "8",  "Indicador": "Ús servei Campanyes puntuals", "Valor": campanyes, "Pct": pct(campanyes)},
        {"#": "10", "Indicador": "Trucades per centraleta",      "Valor": centraleta,"Pct": pct(centraleta)},
    ]


# ─── Export Excel amb dues columnes ──────────────────────────────────────────

def exportar_excel(ind_ajunt: list[dict], ind_general: list[dict],
                   df_raw: pd.DataFrame, municipio: str,
                   d_from: date, d_to: date) -> str:

    safe_name = municipio.replace(" ", "_").replace("/", "-")
    filename  = f"informe_{safe_name}_{d_from.strftime('%Y%m%d')}_{d_to.strftime('%Y%m%d')}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # ── Full "Indicadors" ────────────────────────────────────────────────
        # Construïm el DataFrame combinat: una fila per indicador, 2 columnes de valor
        rows = []
        gen_map = {r["Indicador"]: r for r in ind_general}
        for r in ind_ajunt:
            gen = gen_map.get(r["Indicador"], {})
            rows.append({
                "#":                r["#"],
                "Indicador":        r["Indicador"],
                "Ràtio Ajuntament": r["Valor"],
                "%":                r["Pct"],
                "Ràtio General":    gen.get("Valor", "N/D"),
                "% General":        gen.get("Pct", ""),
            })

        df_ind = pd.DataFrame(rows)
        df_ind.to_excel(writer, sheet_name="Indicadors", index=False)

        # Estil del full Indicadors
        ws = writer.sheets["Indicadors"]
        _estil_indicadors(ws, municipio, d_from, d_to)

        # ── Full "Dades raw" ─────────────────────────────────────────────────
        df_raw.to_excel(writer, sheet_name="Dades raw", index=False)

    return filename


def _parse_numeric(val) -> float | None:
    """
    Extreu un valor numèric comparable d'un valor formatat.
    Exemples: "3.45" → 3.45 | "12.3%" → 12.3 | "14m 37s" → 14.6 | 1234 → 1234.0
    Retorna None si no és comparable (ex: "N/D", "5 municipis").
    """
    if val is None:
        return None
    s = str(val).strip()
    if s in ("N/D", "", "—"):
        return None
    # Percentatge: "12.3%"
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", "."))
        except ValueError:
            return None
    # Temps: "14m 37s"
    if "m" in s and "s" in s:
        try:
            parts = s.replace("s", "").split("m")
            return float(parts[0]) + float(parts[1]) / 60
        except (ValueError, IndexError):
            return None
    # Número pur (pot tenir punts de milers: "1.234")
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


# Indicadors on MENOR valor és millor:
# - Temps mig consulta: menys temps = més eficient
# - Catàleg de tràmits: menys = la gent sap fer els tràmits sola
# - Cita prèvia: menys = menys càrrega presencial
# - Trucades centraleta: menys = menys derivacions, més autonomia
INDICADORS_INVERTITS = {
    "Temps mig consulta",
    "Assistències Catàleg tràmits",
    "Assistències Cita prèvia",
    "Trucades per centraleta",
}


def _estil_indicadors(ws, municipio: str, d_from: date, d_to: date):
    """Aplica estil professional al full Indicadors amb vermell/verd per comparació."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLOR_HEADER  = "1B3A6B"
    COLOR_AJUNT   = "00B4A6"
    COLOR_GENERAL = "4A90D9"
    COLOR_FILA_ALT= "F2F8FF"
    COLOR_SUBIND  = "F7F7F7"
    WHITE         = "FFFFFF"
    GREEN         = "1A7A1A"   # verd fosc per text
    RED           = "C0392B"   # vermell per text
    GREEN_BG      = "E8F5E9"   # fons verd clar
    RED_BG        = "FFEBEE"   # fons vermell clar

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Amplades
    col_widths = [4, 36, 18, 10, 18, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Capçalera (fila 1)
    headers = ["#", "Indicador",
               f"Ràtio {municipio}", "%",
               "Ràtio General", "% General"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    ws.cell(row=1, column=3).fill = PatternFill("solid", fgColor=COLOR_AJUNT)
    ws.cell(row=1, column=4).fill = PatternFill("solid", fgColor=COLOR_AJUNT)
    ws.cell(row=1, column=5).fill = PatternFill("solid", fgColor=COLOR_GENERAL)
    ws.cell(row=1, column=6).fill = PatternFill("solid", fgColor=COLOR_GENERAL)

    # Files de dades
    for row_idx in range(2, ws.max_row + 1):
        num_val  = str(ws.cell(row=row_idx, column=1).value or "")
        ind_val  = str(ws.cell(row=row_idx, column=2).value or "")
        is_sub   = num_val == "2.1"
        is_total = num_val == "—"
        is_alt   = (row_idx % 2 == 0)

        # Compara Ràtio Ajuntament (col 3) vs Ràtio General (col 5)
        val_ajunt   = _parse_numeric(ws.cell(row=row_idx, column=3).value)
        val_general = _parse_numeric(ws.cell(row=row_idx, column=5).value)
        invertit    = ind_val.strip() in INDICADORS_INVERTITS

        color_text = None
        color_bg   = None
        if val_ajunt is not None and val_general is not None and not is_total:
            if val_ajunt > val_general:
                color_text = RED   if invertit else GREEN
                color_bg   = RED_BG if invertit else GREEN_BG
            elif val_ajunt < val_general:
                color_text = GREEN if invertit else RED
                color_bg   = GREEN_BG if invertit else RED_BG

        bg_default = COLOR_SUBIND if is_sub else (COLOR_FILA_ALT if is_alt else WHITE)

        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border    = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col != 2 else "left",
                indent=1 if col == 2 and is_sub else 0
            )

            # Color vermell/verd només a les columnes de valor de l'ajuntament (3 i 4)
            if col in (3, 4) and color_text and not is_sub:
                cell.fill = PatternFill("solid", fgColor=color_bg)
                cell.font = Font(
                    bold=is_total, italic=is_sub, size=9 if is_sub else 10,
                    color=color_text
                )
            else:
                cell.fill = PatternFill("solid", fgColor=bg_default)
                if is_total:
                    cell.font = Font(bold=True, size=10)
                elif is_sub:
                    cell.font = Font(italic=True, color="666666", size=9)
                else:
                    cell.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"


# ─── Multi-període: diverses columnes ────────────────────────────────────────

def exportar_excel_multi_periode(municipio: str,
                                  periodes: list[tuple[date, date]],
                                  ind_per_periode: list[tuple[list, list]]) -> str:
    """
    Genera un Excel amb una columna per cada període.
    ind_per_periode = [(ind_ajunt, ind_general), ...] — un element per període.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    safe_name = municipio.replace(" ", "_").replace("/", "-")
    filename  = f"informe_client_{safe_name}_multiperiode.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicadors"

    # ── Colors ───────────────────────────────────────────────────────────────
    COLOR_HEADER   = "1B3A6B"
    WHITE          = "FFFFFF"
    GREEN          = "1A7A1A"
    RED            = "C0392B"
    GREEN_BG       = "E8F5E9"
    RED_BG         = "FFEBEE"
    FILA_ALT       = "F2F8FF"
    SUBIND_BG      = "F7F7F7"
    PERIOD_COLORS  = ["00B4A6", "2E86AB", "A23B72", "F18F01"]  # un color per període

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Capçaleres: fila 1 = Indicador | Període 1 | Període 2 | ...  ─────────
    labels_periode = [
        f"{d_from.strftime('%d/%m/%y')}–{d_to.strftime('%d/%m/%y')}"
        for d_from, d_to in periodes
    ]

    # Col 1 = #, Col 2 = Indicador, després una col per període
    n_periods = len(periodes)
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=1, column=2, value="Indicador")
    for pi, lbl in enumerate(labels_periode):
        ws.cell(row=1, column=3 + pi, value=lbl)

    # Estil capçalera
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    for pi in range(n_periods):
        ws.column_dimensions[get_column_letter(3 + pi)].width = 20

    for col in range(1, 3 + n_periods):
        cell = ws.cell(row=1, column=col)
        color = PERIOD_COLORS[(col - 3) % len(PERIOD_COLORS)] if col >= 3 else COLOR_HEADER
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 35

    # ── Files de dades ────────────────────────────────────────────────────────
    # Obtenim la llista d'indicadors del primer període com a referència
    ind_base = ind_per_periode[0][0]

    for ri, row_ind in enumerate(ind_base):
        row_idx  = ri + 2
        num_val  = row_ind["#"]
        ind_name = row_ind["Indicador"]
        is_sub   = str(num_val) == "2.1"
        is_total = str(num_val) == "—"
        is_alt   = (row_idx % 2 == 0)

        ws.cell(row=row_idx, column=1, value=num_val)
        ws.cell(row=row_idx, column=2, value=ind_name)

        for pi, (ind_ajunt, ind_general) in enumerate(ind_per_periode):
            col_idx = 3 + pi
            # Troba el valor per aquest indicador en aquest període
            ajunt_row = next((r for r in ind_ajunt   if r["Indicador"] == ind_name), {})
            gen_row   = next((r for r in ind_general if r["Indicador"] == ind_name), {})

            val_ajunt = ajunt_row.get("Valor", "N/D")
            pct_ajunt = ajunt_row.get("Pct", "")
            display   = f"{val_ajunt} {pct_ajunt}".strip() if pct_ajunt else str(val_ajunt)

            cell = ws.cell(row=row_idx, column=col_idx, value=display)

            # Color verd/vermell vs General
            num_a = _parse_numeric(val_ajunt)
            num_g = _parse_numeric(gen_row.get("Valor"))
            invertit = ind_name.strip() in INDICADORS_INVERTITS

            color_text = None
            color_bg   = None
            if num_a is not None and num_g is not None and not is_total and not is_sub:
                if num_a > num_g:
                    color_text = RED   if invertit else GREEN
                    color_bg   = RED_BG if invertit else GREEN_BG
                elif num_a < num_g:
                    color_text = GREEN if invertit else RED
                    color_bg   = GREEN_BG if invertit else RED_BG

            bg_default = SUBIND_BG if is_sub else (FILA_ALT if is_alt else WHITE)
            cell.fill      = PatternFill("solid", fgColor=color_bg or bg_default)
            cell.font      = Font(
                bold=is_total, italic=is_sub,
                size=9 if is_sub else 10,
                color=color_text or ("000000" if not is_sub else "666666")
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        # Estil cols fixes (# i Indicador)
        bg_default = SUBIND_BG if is_sub else (FILA_ALT if is_alt else WHITE)
        for col in (1, 2):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill      = PatternFill("solid", fgColor=bg_default)
            cell.font      = Font(
                bold=is_total, italic=is_sub,
                size=9 if is_sub else 10,
                color="000000" if not is_sub else "666666"
            )
            cell.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center",
                indent=1 if col == 2 and is_sub else 0
            )
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "C2"
    wb.save(filename)
    return filename


def generar_html_multi_periode(municipio: str,
                               periodes: list[tuple[date, date]],
                               ind_per_periode: list[tuple[list, list]]) -> str:
    """
    Genera la taula HTML idèntica a la versió de dos columnes,
    però amb una parella (Ajuntament | General) per cada període.
    """
    INDICADORS_INVERTITS_SET = {
        "Temps mig consulta",
        "Assistències Catàleg tràmits",
        "Assistències Cita prèvia",
        "Trucades per centraleta",
    }

    labels = [
        f"{d_from.strftime('%d/%m/%y')}–{d_to.strftime('%d/%m/%y')}"
        for d_from, d_to in periodes
    ]
    n = len(periodes)
    col_w_ind = 220
    col_w_val = 110

    # Capçalera de grups (una per període)
    group_headers = ""
    for lbl in labels:
        group_headers += f"""
        <th colspan="2" class="period-header">{lbl}</th>"""

    # SubCapçalera (Ajuntament | General per cada període)
    sub_headers = ""
    for _ in labels:
        sub_headers += f"""
        <th class="col-ajunt">{municipio.title()}</th>
        <th class="col-general">Ràtio General</th>"""

    # Files
    ind_base = ind_per_periode[0][0]
    rows_html = ""
    for row_ind in ind_base:
        ind_name = row_ind["Indicador"]
        is_sub   = str(row_ind["#"]) == "2.1"
        is_total = str(row_ind["#"]) == "—"
        row_class = "row-sub" if is_sub else ("row-total" if is_total else "")

        cells = ""
        for pi, (ind_ajunt, ind_general) in enumerate(ind_per_periode):
            ajunt_row = next((r for r in ind_ajunt   if r["Indicador"] == ind_name), {})
            gen_row   = next((r for r in ind_general if r["Indicador"] == ind_name), {})

            val_a = ajunt_row.get("Valor", "N/D")
            pct_a = ajunt_row.get("Pct",   "")
            val_g = gen_row.get("Valor",   "N/D")

            pct_g = gen_row.get("Pct", "")
            display_a = f"{val_a} <span class='pct'>{pct_a}</span>" if pct_a else str(val_a)
            display_g = f"{val_g} <span class='pct'>{pct_g}</span>" if pct_g else str(val_g)

            # Color
            num_a = _parse_numeric(val_a)
            num_g = _parse_numeric(val_g)
            invertit = ind_name.strip() in INDICADORS_INVERTITS_SET
            css_class = ""
            if num_a is not None and num_g is not None and not is_total and not is_sub:
                if num_a > num_g:
                    css_class = "worse" if invertit else "better"
                elif num_a < num_g:
                    css_class = "better" if invertit else "worse"

            sep = ' style="border-left: 2px solid #bbb;"' if pi > 0 else ""
            cells += f'<td class="val {css_class}"{sep}>{display_a}</td>'
            cells += f'<td class="val general">{display_g}</td>'

        rows_html += f"""
        <tr class="{row_class}">
          <td class="ind">{ind_name}</td>
          {cells}
        </tr>"""

    total_width = col_w_ind + n * col_w_val * 2 + 40
    table_width = f"{total_width}px"

    html = f"""<!DOCTYPE html>
<html lang="ca">
<head>
<meta charset="UTF-8">
<title>Dades analítiques — {municipio.title()}</title>
<style>
  body {{
    font-family: 'Segoe UI', Arial, sans-serif;
    background: #f5f5f5;
    display: flex;
    justify-content: center;
    padding: 40px 20px;
  }}
  .wrap {{
    background: white;
    padding: 28px 32px;
    max-width: {table_width};
    width: 100%;
  }}
  .title {{
    font-size: 11px;
    letter-spacing: .12em;
    text-transform: uppercase;
    font-variant: small-caps;
    color: #00A89D;
    margin-bottom: 4px;
  }}
  .subtitle {{
    font-size: 13px;
    color: #555;
    margin-bottom: 18px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
  }}
  th {{
    font-size: 11px;
    font-weight: 600;
    padding: 7px 10px;
    text-align: center;
    border-bottom: 2px solid #ccc;
    color: #444;
    background: #f9f9f9;
  }}
  th.period-header {{
    font-size: 11px;
    color: #333;
    background: #f0f0f0;
    border-bottom: 1px solid #ddd;
    border-left: 2px solid #bbb;
    padding: 5px 10px;
  }}
  th.period-header:first-of-type {{
    border-left: none;
  }}
  th.col-ajunt {{
    color: #00A89D;
    background: #f9f9f9;
    border-left: 2px solid #bbb;
    min-width: {col_w_val}px;
  }}
  th.col-general {{
    color: #777;
    background: #f9f9f9;
    min-width: {col_w_val}px;
  }}
  td.ind {{
    font-size: 12px;
    color: #333;
    padding: 7px 10px;
    border-bottom: 1px solid #eee;
    min-width: {col_w_ind}px;
  }}
  td.val {{
    font-size: 12px;
    font-weight: 700;
    text-align: center;
    padding: 7px 10px;
    border-bottom: 1px solid #eee;
  }}
  td.general {{
    color: #888;
    font-weight: 400;
  }}
  .better {{ color: #00A89D; }}
  .worse  {{ color: #D9534F; }}
  .pct    {{ font-size: 10px; font-weight: 400; color: #999; }}
  tr.row-sub td {{ font-size: 11px; color: #777; font-weight: 400; }}
  tr.row-sub td.val {{ font-weight: 400; }}
  tr.row-total td {{ font-weight: 700; border-top: 2px solid #ddd; }}
  tr:hover td {{ background: #fafafa; }}
  .download {{
    margin-top: 16px;
    text-align: right;
  }}
  .download a {{
    font-size: 11px;
    color: #00A89D;
    text-decoration: none;
    border: 1px solid #00A89D;
    padding: 4px 10px;
    border-radius: 3px;
  }}
  .download a:hover {{ background: #00A89D; color: white; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="title">Dades analítiques principals</div>
  <div class="subtitle">{municipio.title()} — comparativa anual</div>
  <table>
    <thead>
      <tr>
        <th rowspan="2" style="text-align:left; min-width:{col_w_ind}px;">Indicador</th>
        {group_headers}
      </tr>
      <tr>
        {sub_headers}
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
</div>
</body>
</html>"""
    return html


def servir_html(html: str, port: int = 8765):
    """Escriu l'HTML a /tmp i el serveix a localhost."""
    import os, threading, http.server
    os.makedirs("/tmp/informe-taula", exist_ok=True)
    with open("/tmp/informe-taula/index.html", "w", encoding="utf-8") as f:
        f.write(html)

    class Handler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *a, **kw):
            super().__init__(*a, directory="/tmp/informe-taula", **kw)
        def log_message(self, *_): pass

    def run():
        with http.server.HTTPServer(("", port), Handler) as srv:
            srv.serve_forever()

    t = threading.Thread(target=run, daemon=True)
    t.start()
    print(f"  🌐 Taula disponible a: http://localhost:{port}")
    import webbrowser
    webbrowser.open(f"http://localhost:{port}")


def download_full_cache(client: AdtendeClient, d_from: date, d_to: date,
                        municipio: str = None) -> pd.DataFrame:
    """
    Descarrega tots els mesos del rang d'una sola vegada (sense filtre de dates).
    Retorna el DataFrame complet sense aplicar filtre de rang — es farà client-side per cada període.
    """
    mesos = months_in_range(d_from, d_to)
    dfs = []
    label = municipio if municipio else "GENERAL"
    total_mesos = len(mesos)

    for i, (y, m) in enumerate(mesos):
        m_from, m_to = month_window(y, m)
        filters = [{"type": "date", "variable": "td_managed",
                    "values": {"gte": m_from, "lt": m_to}}]
        try:
            df_m = client.query("tickets_enriquits", filters=filters)
            if municipio:
                sub = df_m[
                    (df_m["des_client"]  == municipio) &
                    (df_m["des_project"] == "OAC 360º")
                ].copy()
            else:
                sub = df_m[df_m["des_project"] == "OAC 360º"].copy()
            dfs.append(sub)
            print(f"  [{label}] {y}-{m:02d} ({i+1}/{total_mesos}) → {len(sub):>5} tickets")
        except Exception as e:
            print(f"  [{label}] {y}-{m:02d} → ERROR: {e}")

    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def filter_period(df_cache: pd.DataFrame, d_from: date, d_to: date) -> pd.DataFrame:
    """Aplica el filtre doble (td_managed + td_created) sobre el cache ja descarregat."""
    s_from = str(d_from)
    s_to   = str(d_to + timedelta(days=1))
    return df_cache[
        (df_cache["td_managed"] >= s_from) & (df_cache["td_managed"] < s_to) &
        (df_cache["td_created"] >= s_from) & (df_cache["td_created"] < s_to)
    ].copy().reset_index(drop=True)


def generate_multi_period_report(municipio: str, periodes: list[tuple[date, date]]):
    """
    Genera l'informe de client per a múltiples períodes.
    Descarrega UNA SOLA VEGADA tot el rang global i filtra client-side.
    """
    client = AdtendeClient()
    print("Autenticant...")
    client.login()

    # Rang global: del mínim al màxim de tots els períodes
    global_from = min(d for d, _ in periodes)
    global_to   = max(d for _, d in periodes)

    print(f"\n  Rang global: {global_from.strftime('%d/%m/%Y')} → {global_to.strftime('%d/%m/%Y')}")
    mesos_total = len(months_in_range(global_from, global_to))
    print(f"  Descarregant {mesos_total} mesos per {municipio} (1 descàrrega)...")
    df_ajunt_cache = download_full_cache(client, global_from, global_to, municipio=municipio)

    print(f"\n  Descarregant {mesos_total} mesos GENERAL (1 descàrrega)...")
    df_general_cache = download_full_cache(client, global_from, global_to, municipio=None)

    # Calcula indicadors per cada període filtrant el cache
    print(f"\n  Calculant indicadors per cada període...")
    ind_per_periode = []
    for d_from, d_to in periodes:
        df_a = filter_period(df_ajunt_cache,   d_from, d_to)
        df_g = filter_period(df_general_cache, d_from, d_to)
        ind_a = calcular_indicadors(df_a, es_general=False)
        ind_g = calcular_indicadors(df_g, es_general=True)
        ind_per_periode.append((ind_a, ind_g))
        lbl = f"{d_from.strftime('%d/%m/%y')}–{d_to.strftime('%d/%m/%y')}"
        print(f"    {lbl}: {len(df_a)} tickets {municipio} | {len(df_g)} tickets GENERAL")

    print(f"\n{'─'*60}")
    print("  Generant Excel...")
    filename = exportar_excel_multi_periode(municipio, periodes, ind_per_periode)
    print(f"  ✅ Excel: {filename}")

    print("  Generant taula HTML...")
    html = generar_html_multi_periode(municipio, periodes, ind_per_periode)
    servir_html(html)

    return filename


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("\n╔══════════════════════════════════════════════════════╗")
    print("║     Extracció dades analítiques — OAC 360            ║")
    print("║     Ràtio Ajuntament + Ràtio General                 ║")
    print("╚══════════════════════════════════════════════════════╝\n")

    # ── Inputs ──────────────────────────────────────────────────────────────
    municipio = input("Municipi (tal com apareix al BI): ").strip()
    if not municipio:
        print("Error: cal introduir un municipi.")
        sys.exit(1)

    while True:
        try:
            d_from = parse_date(input("Data inici  (DD/MM/AAAA): "))
            break
        except (ValueError, IndexError):
            print("  ↳ Format incorrecte. Exemple: 06/10/2021")

    while True:
        try:
            d_to = parse_date(input("Data fi     (DD/MM/AAAA): "))
            break
        except (ValueError, IndexError):
            print("  ↳ Format incorrecte. Exemple: 06/10/2022")

    if d_to < d_from:
        print("Error: la data fi ha de ser posterior a la data inici.")
        sys.exit(1)

    print(f"\n  Municipi : {municipio}")
    print(f"  Període  : {d_from.strftime('%d/%m/%Y')} → {d_to.strftime('%d/%m/%Y')}")

    client = AdtendeClient()
    client.login()

    # ── Descàrrega Ajuntament ────────────────────────────────────────────────
    print(f"\n  Descarregant dades del municipi...\n")
    df_ajunt = download_and_filter(client, d_from, d_to, municipio=municipio)

    if df_ajunt.empty:
        print("\n⚠️  No s'han trobat dades. Comprova el nom del municipi.")
        sys.exit(1)

    # ── Descàrrega General ───────────────────────────────────────────────────
    print(f"\n  Descarregant dades generals (tots els municipis)...\n")
    df_general = download_and_filter(client, d_from, d_to, municipio=None)

    # ── Càlcul ──────────────────────────────────────────────────────────────
    ind_ajunt   = calcular_indicadors(df_ajunt,   es_general=False)
    ind_general = calcular_indicadors(df_general, es_general=True)

    # ── Mostra resultats ────────────────────────────────────────────────────
    sep = "─" * 72
    print(f"\n  {sep}")
    print(f"  {'Indicador':<32} {'Ajuntament':>14}  {'General':>14}")
    print(f"  {sep}")
    gen_map = {r["Indicador"]: r for r in ind_general}
    for r in ind_ajunt:
        gen = gen_map.get(r["Indicador"], {})
        val_a = f"{r['Valor']} {r['Pct']}".strip()
        val_g = f"{gen.get('Valor','N/D')} {gen.get('Pct','')}".strip()
        print(f"  {r['Indicador']:<32} {val_a:>14}  {val_g:>14}")
    print(f"  {sep}\n")

    # ── Exporta ─────────────────────────────────────────────────────────────
    filename = exportar_excel(ind_ajunt, ind_general, df_ajunt, municipio, d_from, d_to)
    print(f"  ✅  Exportat a: {filename}\n")


if __name__ == "__main__":
    main()
